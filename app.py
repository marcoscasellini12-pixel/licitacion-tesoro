import streamlit as st
import io, re, base64
from datetime import datetime
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage
import os

st.set_page_config(page_title="Licitación del Tesoro", page_icon="🏦", layout="centered")
st.markdown("""<style>
.stButton>button{background:#1a3a5c;color:white;font-weight:bold;width:100%;padding:.6rem;font-size:1.05rem;border-radius:6px;border:none;}
.stButton>button:hover{background:#e26b0a;}
</style>""", unsafe_allow_html=True)

NARANJA="FFE26B0A"; BLANCO="FFFFFFFF"; NEGRO="FF000000"
GRIS_OSC="FFBFBFBF"; GRIS_CLA="FFD8D8D8"
SZ=11

MESES={"ENERO":1,"FEBRERO":2,"MARZO":3,"ABRIL":4,"MAYO":5,"JUNIO":6,"JULIO":7,"AGOSTO":8,"SEPTIEMBRE":9,"OCTUBRE":10,"NOVIEMBRE":11,"DICIEMBRE":12}
MESES_ES={1:"enero",2:"febrero",3:"marzo",4:"abril",5:"mayo",6:"junio",7:"julio",8:"agosto",9:"septiembre",10:"octubre",11:"noviembre",12:"diciembre"}

def pf(t):
    if not t: return None
    m=re.search(r"(\d{1,2})\s+DE\s+(\w+)\s+(?:DE\s+)?(\d{4})",str(t).upper())
    if not m: return None
    mes=MESES.get(m.group(2))
    if not mes: return None
    try: return datetime(int(m.group(3)),mes,int(m.group(1)))
    except: return None

def pddmm(t):
    m=re.search(r"(\d{2})/(\d{2})/(\d{4})",str(t))
    if m:
        try: return datetime(int(m.group(3)),int(m.group(2)),int(m.group(1)))
        except: pass
    return None

def fstr(dt):
    if not dt: return ""
    return f"{dt.day} de {MESES_ES[dt.month]} de {dt.year}"

def dias_entre(venc, emision):
    if not venc or not emision: return ""
    d = (venc - emision).days
    if d > 365:
        a = d // 365
        meses_resto = (d % 365) // 30  # truncar, no redondear
        if meses_resto > 0:
            return f"Aprox. {a} año{'s' if a>1 else ''} y {meses_resto} meses"
        return f"Aprox. {a} año{'s' if a>1 else ''}"
    return f"{d} días"

def ticker_fecha(T,ticker):
    idx=T.find(ticker)
    if idx<0: return None
    frag=T[max(0,idx-500):idx+500]
    m=re.search(r"(\d{2}/\d{2}/\d{4})",frag)
    if m: return pddmm(m.group(1))
    m2=re.search(r"VENCIMIENTO\s+([\d\w\s]+?)\s*\("+ticker,T)
    if m2: return pf(m2.group(1))
    return None

# ── Estilos ───────────────────────────────────────────────────────────────────
def F(rgb): return PatternFill("solid",start_color=rgb)
def fn(bold=False,color=NEGRO): return Font(bold=bold,size=SZ,color=color,name="Calibri")
def al(h="center",v="center",wrap=True): return Alignment(horizontal=h,vertical=v,wrap_text=wrap)
def B():
    s=Side(style="medium",color=BLANCO)
    return Border(left=s,right=s,top=s,bottom=s)

def aplicar(cell,rgb,valor=None,bold=False,color=NEGRO,h="center",wrap=True,fmt=None):
    if valor is not None: cell.value=valor
    cell.fill=F(rgb); cell.font=fn(bold=bold,color=color)
    cell.alignment=al(h=h,wrap=wrap); cell.border=B()
    if fmt: cell.number_format=fmt

def aplicar_rango(ws,r,c1,c2,rgb):
    for c in range(c1,c2+1): aplicar(ws.cell(row=r,column=c),rgb)

def merge_escribir(ws,r,c1,c2,rgb,valor,h="center"):
    aplicar_rango(ws,r,c1,c2,rgb)
    cell=ws.cell(row=r,column=c1,value=valor)
    cell.font=fn(); cell.alignment=al(h=h); cell.border=B()
    if c2>c1: ws.merge_cells(start_row=r,start_column=c1,end_row=r,end_column=c2)

# ── Extracción PDF ────────────────────────────────────────────────────────────
def extraer_pdf(pdf_bytes):
    txt=""
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for p in pdf.pages: txt+=(p.extract_text() or "")+"\n"
    T=re.sub(r"\s+"," ",txt.upper()).strip()

    m_liq=re.search(r"LIQUIDACI[ÓO]N.*?(\d{1,2}\s+DE\s+\w+\s+DE\s+\d{4})",T)
    fecha_liq=pf(m_liq.group(1)) if m_liq else None
    if not fecha_liq:
        m2=re.search(r"LIQUIDACI[ÓO]N.*?(\d{2}/\d{2}/\d{4})",T)
        if m2: fecha_liq=pddmm(m2.group(1))
    fecha_em=fecha_liq

    mh=re.search(r"(\d{1,2}:\d{2})\s+HORAS.*?(\d{1,2}:\d{2})\s+HORAS",T)
    h_ini=mh.group(1) if mh else "10:00"; h_fin=mh.group(2) if mh else "15:00"
    mf=re.search(r"(?:día|dia)\s+\w+\s+(\d{1,2}\s+de\s+\w+\s+de\s+\d{4})",txt,re.IGNORECASE)
    fecha_lic_str=mf.group(1).strip() if mf else ""
    sv=""
    msv=re.search(r"segunda vuelta[^\n]*?(\d{1,2}:\d{2}).*?(\d{1,2}:\d{2}).*?(\d{1,2}\s+de\s+\w+\s+de\s+\d{4})",txt,re.IGNORECASE|re.DOTALL)
    if msv: sv=f"Segunda Vuelta BONAR 2027: desde las {msv.group(1)} hs hasta las {msv.group(2)} hs del  {msv.group(3).strip()}"

    sus_ll='Pesos al tipo de cambio de referencia publicado por el BCRA (Comunicación "A" 3500) correspondiente al día hábil previo a la fecha de licitación (T-1)'
    pf_,pv_,cer_,usd_=[],[],[],[]
    seen=set()
    def add(l,k,d):
        if k not in seen: seen.add(k); l.append(d)

    def inst_pesos(label,f,tasa,precio,ajuste,parametro):
        return {"label":label,"vencimiento":f,"tasa":tasa,"precio":precio,"ajuste":ajuste,
                "parametro":parametro,"amort":"Íntegra al vencimiento",
                "monto":"Hasta el monto máximo autorizado por la normativa vigente"}

    # LECAP nuevas
    for m in re.finditer(r"LETRA DEL TESORO NACIONAL CAPITALIZABLE EN PESOS CON VENCIMIENTO ([\d\w\s]+?)\(NUEVA\)",T):
        f=pf(m.group(1))
        if f: add(pf_,f"LN{f}",inst_pesos("LECAP (nueva)",f,"A licitar","$ 1.000,00 por cada VNO $ 1.000","N/A","TEM"))
    # LECAP reaperturas
    for m in re.finditer(r"LETRA DEL TESORO NACIONAL CAPITALIZABLE EN PESOS CON VENCIMIENTO [\d\w\s]+?\((\w+)\s*-\s*REAPERTURA\)",T):
        tk=m.group(1); f=ticker_fecha(T,tk)
        add(pf_,f"LR{tk}",inst_pesos(f"LECAP ({tk} - reapertura)",f,"A licitar","A licitar","N/A","Precio"))
    # BONCAP reaperturas
    for m in re.finditer(r"BONO DEL TESORO NACIONAL CAPITALIZABLE EN PESOS CON VENCIMIENTO [\d\w\s]+?\((\w+)\s*-\s*REAPERTURA\)",T):
        tk=m.group(1); f=ticker_fecha(T,tk)
        add(pf_,f"BCAP{tk}",inst_pesos(f"BONCAP ({tk} - reapertura)",f,"A licitar","A licitar","N/A","Precio"))
    # LETAMAR reaperturas
    for m in re.finditer(r"LETRA DEL TESORO NACIONAL EN PESOS A TASA TAMAR CON VENCIMIENTO ([\d\w\s]+?)\((\w+)\s*-\s*REAPERTURA\)",T):
        f=pf(m.group(1)); tk=m.group(2)
        add(pv_,f"LT{tk}",inst_pesos(f"LETAM ({tk} - reapertura)",f,"","A licitar","N/A","Precio"))
    # BOTAMAR reaperturas
    for m in re.finditer(r"BONO DEL TESORO NACIONAL EN PESOS A TASA TAMAR CON VENCIMIENTO ([\d\w\s]+?)\((\w+)\s*[-–]\s*REAPERTURA\)",T):
        f=pf(m.group(1)); tk=m.group(2)
        add(pv_,f"BT{tk}",inst_pesos(f"BOTAM ({tk} - reapertura)",f,"","A licitar","N/A","Precio"))
    # BOTAMAR nuevos
    for m in re.finditer(r"BONO DEL TESORO NACIONAL EN PESOS A TASA TAMAR CON VENCIMIENTO ([\d\w\s]+?)\(NUEVO\)",T):
        f=pf(m.group(1))
        if f: add(pv_,f"BTN{f}",inst_pesos("BOTAM (nuevo)",f,"A licitar","$ 1.000,00 por cada VNO $ 1.000","N/A","Tasa"))
    # LECER reaperturas
    for m in re.finditer(r"LETRA DEL TESORO NACIONAL EN PESOS AJUSTADA? POR CER A DESCUENTO VENCIMIENTO [\d\w\s]+?\((\w+)\s*-\s*REAPERTURA\)",T):
        tk=m.group(1); f=ticker_fecha(T,tk)
        add(cer_,f"CR{tk}",inst_pesos(f"LECER ({tk} - reapertura)",f,"Cero Cupón","A licitar","CER","Precio"))
    # LECER nuevas
    for m in re.finditer(r"LETRA DEL TESORO NACIONAL EN PESOS AJUSTADA? POR CER A DESCUENTO VENCIMIENTO ([\d\w\s]+?)\(NUEVA\)",T):
        f=pf(m.group(1))
        if f: add(cer_,f"CN{f}",inst_pesos("LECER (Nueva)",f,"Cero Cupón","A licitar","CER","Precio"))
    # BONCER reaperturas
    for m in re.finditer(r"BONO DEL TESORO NACIONAL EN PESOS CERO CUP[ÓO]N CON AJUSTE POR CER VENCIMIENTO [\d\w\s]+?\((\w+)\s*[-–]\s*REAPERTURA\)",T):
        tk=m.group(1); f=ticker_fecha(T,tk)
        add(cer_,f"BR{tk}",inst_pesos(f"BONCER ({tk} – reapertura)",f,"Cero Cupón","A licitar","CER","Precio"))
    # BONCER nuevos
    for m in re.finditer(r"BONO DEL TESORO NACIONAL EN PESOS CERO CUP[ÓO]N CON AJUSTE POR CER VENCIMIENTO ([\d\w\s]+?)\(NUEVO\)",T):
        f=pf(m.group(1))
        if f: add(cer_,f"BRN{f}",inst_pesos("BONCER (nuevo)",f,"Cero Cupón","A licitar","CER","Precio"))
    # LELINK reaperturas
    for m in re.finditer(r"LETRA DEL TESORO NACIONAL VINCULADA AL D[ÓO]LAR ESTADOUNIDENSE CERO CUP[ÓO]N CON VENCIMIENTO [\d\w\s]+?\((\w+)\s*-\s*REAPERTURA\)",T):
        tk=m.group(1); f=ticker_fecha(T,tk)
        add(usd_,f"LL{tk}",{"tipo":"LELINK","label":f"LELINK ({tk} - reapertura)","vencimiento":f,"tasa":"Cero Cupón","precio":"A licitar","parametro":"Precio","mon_em":"Dólares Estadounidenses","mon_sus":sus_ll,"mon_pago":"Pesos al tipo de cambio aplicable","amort":"Íntegra al vencimiento","monto":"Hasta el monto máximo autorizado por la normativa vigente"})
    # LELINK nuevas
    for m in re.finditer(r"LETRA DEL TESORO NACIONAL VINCULADA AL D[ÓO]LAR ESTADOUNIDENSE CERO CUP[ÓO]N CON VENCIMIENTO ([\d\w\s]+?)\(NUEVA\)",T):
        f=pf(m.group(1))
        if f: add(usd_,f"LLN{f}",{"tipo":"LELINK","label":"LELINK (Nuevo)","vencimiento":f,"tasa":"Cero Cupón","precio":"A licitar","parametro":"Precio","mon_em":"Dólares Estadounidenses","mon_sus":sus_ll,"mon_pago":"Pesos al tipo de cambio aplicable","amort":"Íntegra al vencimiento","monto":"Hasta el monto máximo autorizado por la normativa vigente"})
    # BONAR
    for m in re.finditer(r"BONO DEL TESORO NACIONAL EN DOLARES ESTADOUNIDENSES.*?VENCIMIENTO [\d\w\s]+?\((\w+)\s*-\s*REAPERTURA\)",T):
        tk=m.group(1); f=ticker_fecha(T,tk)
        mmx=re.search(r"USD\s*(\d+)\s*MILLONES.*?(?:EN\s+)?(?:LA\s+)?PRIMERA VUELTA",T)
        monto_str=f"USD {mmx.group(1)} Millones en primera vuelta. (*)" if mmx else ""
        m_tna=re.search(r"ESTADOUNIDENSES\s+(\d+%)",T); tna=m_tna.group(1) if m_tna else "6%"
        mr=re.search(r"RESCATE ANTICIPADO.*?(\d{1,2}\s+DE\s+\w+\s+DE\s+\d{4})",T)
        fr_str=fstr(pf(mr.group(1))) if mr else ""
        add(usd_,f"BN{tk}",{"tipo":"BONAR","label":f"BONAR 2027 ({tk} - reapertura)","vencimiento":f,"tasa":f"TNA {tna} a pagar mensualmente","precio":"A licitar","parametro":"Precio","mon_em":"Dólares Estadounidenses","mon_sus":"En Dólares Estadounidenses","mon_pago":"En Dólares Estadounidenses","amort":"Íntegra al vencimiento","monto":monto_str,"opcion":"Los tenedores de los Bonos podrán ejercer, por única vez, una opción de rescate anticipado, total o parcial, del capital del Bono.","fecha_opcion":fr_str,"pago_int":"Los intereses serán pagaderos en Pesos por semestre vencido los días 30 de mayo y 30 de noviembre de cada año hasta la fecha de vencimiento"})

    # ── Canje (si existe) ────────────────────────────────────────────────────
    canje = []
    titulo_elegible = ""
    liq_canje = None
    m_canje_ini = re.search(r"CONVERSI[ÓO]N DEL (\w+)", T)
    if m_canje_ini:
        m_fin_list = list(re.finditer(r"BUENOS AIRES", T))
        sec_canje = T[m_canje_ini.start():m_fin_list[-1].start()] if m_fin_list else ""
        # Título elegible
        m_te = re.search(r"(\w+)\s+(\d{2}/\d{2}/\d{4})", sec_canje)
        titulo_elegible = m_canje_ini.group(1) if m_canje_ini else ""
        # Liq canje T+3
        m_liq_c = re.search(r"(\d{1,2}\s+DE\s+\w+\s+DE\s+\d{4})\s*\(T\+3\)", sec_canje)
        liq_canje = pf(m_liq_c.group(1)) if m_liq_c else None
        # Opciones
        pos_op = [(m.start(), int(m.group(1))) for m in re.finditer(r"OPCI[ÓO]N\s+(\d)\)", sec_canje)]
        for idx, (start, num) in enumerate(pos_op):
            end = pos_op[idx+1][0] if idx+1 < len(pos_op) else len(sec_canje)
            bloque = sec_canje[start:end]
            m_f = re.search(r"(\d{1,2}/\d{1,2}/\d{4})", bloque)
            f = pddmm(m_f.group(1)) if m_f else None
            if "TAMAR" in bloque and "TASA TAMAR" in bloque:
                canje.append({"label":"BOTAM (nuevo)","vencimiento":f,"tasa":"A licitar",
                    "precio":"$ 1.000,00 por cada VNO $ 1.000","ajuste":"N/A","parametro":"Tasa",
                    "amort":"Íntegra al vencimiento","monto":"Hasta el monto máximo autorizado por la normativa vigente"})
            elif "CER" in bloque and "AJUSTE" in bloque:
                canje.append({"label":"BONCER (nuevo)","vencimiento":f,"tasa":"Cero Cupón",
                    "precio":"A licitar","ajuste":"CER","parametro":"Precio",
                    "amort":"Íntegra al vencimiento","monto":"Hasta el monto máximo autorizado por la normativa vigente"})

    return {"pf":pf_,"pv":pv_,"cer":cer_,"usd":usd_,
            "canje":canje,"titulo_elegible":titulo_elegible,"liq_canje":liq_canje,"liq_canje_str":fstr(liq_canje),
            "h":{"h_ini":h_ini,"h_fin":h_fin,"fecha_lic":fecha_lic_str,"sv":sv,
                 "liq":fecha_liq,"liq_str":fstr(fecha_liq),"em":fecha_em}}

# ── Logos (cargados desde archivos adjuntos en el repo o desde bytes subidos) ──
LOGO_HIP_PATH = os.path.join(os.path.dirname(__file__), "logo_hipotecario.png")
LOGO_MIN_PATH = os.path.join(os.path.dirname(__file__), "logo_ministerio.png")

def redimensionar_logo(path, alto_px):
    """Redimensiona manteniendo aspecto, devuelve BytesIO PNG."""
    img = PILImage.open(path)
    w, h = img.size
    nuevo_w = int(w * alto_px / h)
    img = img.resize((nuevo_w, alto_px), PILImage.LANCZOS)
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    return buf, nuevo_w, alto_px

# ── Excel ─────────────────────────────────────────────────────────────────────
def generar_excel(datos):
    wb=Workbook(); ws=wb.active
    h=datos["h"]; fecha_em=h["em"]
    if h["liq"]: ws.title=h["liq"].strftime("%d.%m.%Y")
    ws.sheet_view.showGridLines=False

    ws.column_dimensions["A"].width=4
    ws.column_dimensions["B"].width=4
    ws.column_dimensions["C"].width=30
    for i in range(4,18): ws.column_dimensions[get_column_letter(i)].width=24

    # Alturas de fila para logos
    ws.row_dimensions[1].height=15
    ws.row_dimensions[2].height=50  # fila con logos
    ws.row_dimensions[3].height=10
    ws.row_dimensions[4].height=15

    # ── Insertar logos ────────────────────────────────────────────────────────
    ALTO_LOGO = 50  # puntos ≈ px en pantalla

    try:
        buf_hip, w_hip, _ = redimensionar_logo(LOGO_HIP_PATH, ALTO_LOGO)
        img_hip = XLImage(buf_hip)
        img_hip.anchor = "C2"
        ws.add_image(img_hip)
    except Exception:
        pass

    try:
        buf_min, w_min, _ = redimensionar_logo(LOGO_MIN_PATH, ALTO_LOGO)
        img_min = XLImage(buf_min)
        img_min.anchor = "G2"
        ws.add_image(img_min)
    except Exception:
        pass

    CL=3; CD=4

    # ── Header ────────────────────────────────────────────────────────────────
    # Todo en una sola celda combinada — negrita hasta ":" resto normal
    # openpyxl no soporta rich text, entonces ponemos la línea completa en negrita
    # para la parte bold y en celda aparte el texto normal, pero SIN separación visual
    # La solución más limpia: una sola celda con texto completo, negrita en label

    def hdr_linea(r, negrita, normal, h_row=18):
        ws.row_dimensions[r].height = h_row
        # Celda principal con texto COMPLETO combinada hasta col 12
        cell = ws.cell(row=r, column=CL, value=negrita + normal)
        # Aplicar negrita a toda la celda (openpyxl no soporta rich text parcial)
        cell.font = Font(bold=False, size=SZ, name="Calibri")
        cell.alignment = al(h="left", wrap=False)
        ws.merge_cells(start_row=r, start_column=CL, end_row=r, end_column=12)
        # Sobreescribir con dos runs: no es posible en openpyxl directamente
        # Solución: poner la parte bold en una celda angosta y la normal en la siguiente
        # Pero esto los separa visualmente. Mejor: texto completo sin bold diferenciado.
        # Para respetar el formato: bold en toda la celda para las líneas de label
        cell.font = Font(bold=True, size=SZ, name="Calibri")

    ws.row_dimensions[7].height=18
    c=ws.cell(row=7,column=CL,value="LICITACIÓN DEL TESORO")
    c.font=Font(bold=True,size=SZ,name="Calibri"); c.alignment=al(h="left",wrap=False)

    ws.row_dimensions[9].height=18
    c=ws.cell(row=9,column=CL,value="LICITACION POR EFECTIVO")
    c.font=Font(bold=True,size=SZ,name="Calibri"); c.alignment=al(h="left",wrap=False)

    # Período — "Período de Licitación Pública:" bold + resto normal en UNA celda combinada
    def hdr_bicolor(r, parte_bold, parte_normal, h_row=18):
        ws.row_dimensions[r].height = h_row
        # Celda bold (col C, angosta)
        c1 = ws.cell(row=r, column=CL, value=parte_bold)
        c1.font = Font(bold=True, size=SZ, name="Calibri")
        c1.alignment = al(h="left", wrap=False)
        # Celda normal (col D en adelante, combinada)
        c2 = ws.cell(row=r, column=CL+1, value=parte_normal)
        c2.font = Font(bold=False, size=SZ, name="Calibri")
        c2.alignment = al(h="left", wrap=False)
        ws.merge_cells(start_row=r, start_column=CL+1, end_row=r, end_column=12)

    hdr_bicolor(11, "Período de Licitación Pública:",
                f" desde las {h['h_ini']} hs hasta las {h['h_fin']} hs del  {h['fecha_lic']}")
    if h["sv"]:
        partes = h["sv"].split(":",1)
        hdr_bicolor(12, partes[0]+":", partes[1] if len(partes)>1 else "")
    hdr_bicolor(13, "Fecha de Liquidación:", f" {h['liq_str']} (T+2)")

    fila=[15]
    def R(n=1): r=fila[0]; fila[0]+=n; return r

    def bloque(titulo, insts, es_usd=False):
        if not insts: return
        n=len(insts); c1=CD; c2=CD+n-1

        r=R(); ws.row_dimensions[r].height=21
        ct=ws.cell(row=r,column=CL,value=titulo)
        ct.font=Font(italic=True,bold=True,size=SZ,name="Calibri")
        ct.alignment=al(h="left",wrap=False)
        R()

        r=R(); ws.row_dimensions[r].height=42
        cc=ws.cell(row=r,column=CL); cc.fill=F(NARANJA); cc.border=B()
        for i,inst in enumerate(insts):
            c=ws.cell(row=r,column=c1+i,value=inst["label"])
            c.fill=F(NARANJA); c.font=fn(color=BLANCO)
            c.alignment=al(); c.border=B()

        def fila_ind(label, getter, gris, h_row=21, fmt=None):
            r=R(); ws.row_dimensions[r].height=h_row
            aplicar(ws.cell(row=r,column=CL),gris,valor=label,h="left")
            for i,inst in enumerate(insts):
                val=getter(inst)
                c=ws.cell(row=r,column=c1+i)
                aplicar(c,gris,valor=val if val is not None else "")
                if fmt: c.number_format=fmt

        def fila_merge(label, valor, gris, h_row=21):
            r=R(); ws.row_dimensions[r].height=h_row
            aplicar(ws.cell(row=r,column=CL),gris,valor=label,h="left")
            merge_escribir(ws,r,c1,c2,gris,valor)

        if not es_usd:
            fila_ind("Vencimiento",     lambda x:x["vencimiento"], GRIS_CLA, fmt="D/M/YYYY")
            fila_ind("Plazo",           lambda x:dias_entre(x["vencimiento"],fecha_em), GRIS_OSC)
            fila_merge("Moneda de emision",     "Pesos", GRIS_CLA)
            fila_merge("Moneda de Suscripcion", "Pesos", GRIS_CLA)
            fila_merge("Moneda de Pago",        "Pesos", GRIS_CLA)
            fila_ind("Tasa de interés ",  lambda x:x.get("tasa",""),     GRIS_OSC)
            fila_ind("Precio",            lambda x:x.get("precio",""),   GRIS_CLA, h_row=42)
            fila_ind("Ajuste de capital", lambda x:x.get("ajuste",""),   GRIS_OSC)
            fila_ind("Párametro a licitar",lambda x:x.get("parametro",""),GRIS_CLA)
            fila_merge("Amortización",           "Íntegra al vencimiento", GRIS_OSC)
            fila_merge("Monto Máximo a Licitar", "Hasta el monto máximo autorizado por la normativa vigente", GRIS_CLA)
            fila_merge("Ley aplicable",          "Ley de la REPÚBLICA ARGENTINA", GRIS_OSC)
        else:
            fila_ind("Vencimiento",       lambda x:x["vencimiento"],      GRIS_CLA, fmt="D/M/YYYY")
            fila_ind("Moneda de emision", lambda x:x.get("mon_em",""),    GRIS_OSC)
            fila_ind("Moneda de Suscripcion",lambda x:x.get("mon_sus",""),GRIS_CLA,h_row=84)
            fila_ind("Moneda de Pago",    lambda x:x.get("mon_pago",""),  GRIS_OSC, h_row=42)
            fila_ind("Tasa de interés",   lambda x:x.get("tasa",""),      GRIS_CLA, h_row=42)
            fila_ind("Precio",            lambda x:x.get("precio",""),    GRIS_OSC)
            fila_ind("Párametro a licitar",lambda x:x.get("parametro",""),GRIS_CLA)
            fila_ind("Amortización",      lambda x:x.get("amort",""),     GRIS_OSC)
            fila_ind("Monto Máximo a Licitar",lambda x:x.get("monto",""), GRIS_CLA, h_row=42)
            fila_merge("Ley aplicable",   "Ley de la REPÚBLICA ARGENTINA", GRIS_OSC)

        R()

    def bonar_especial(usd_list):
        bonar=[x for x in usd_list if x.get("tipo")=="BONAR"]
        if not bonar: return
        b=bonar[0]; n=len(usd_list); c1=CD; c2=CD+n-1
        r=R(); ws.row_dimensions[r].height=21
        ct=ws.cell(row=r,column=CL,value="Instrumentros a licitar en dólares")
        ct.font=Font(italic=True,bold=True,size=SZ,name="Calibri"); ct.alignment=al(h="left",wrap=False)
        for label,val,h_row,gris in [
            ("Opción de rescate anticipado",    b.get("opcion",""),    42,GRIS_CLA),
            ("Fecha de ejercicio de la Opción", b.get("fecha_opcion",""),21,GRIS_OSC),
            ("Forma de pago de los servicios\n de interés",b.get("pago_int",""),63,GRIS_CLA),
        ]:
            r=R(); ws.row_dimensions[r].height=h_row
            aplicar(ws.cell(row=r,column=CL),gris,valor=label,h="left")
            merge_escribir(ws,r,c1,c2,gris,val,h="left")
        R()

    pesos=datos["pf"]+datos["pv"]
    if pesos:        bloque("Instrumentos a licitar en pesos a tasa fija y tasa variable:",pesos)
    if datos["cer"]: bloque("Instrumentos a licitar en pesos ajustados por CER:",datos["cer"])
    if datos["usd"]:
        bonar_especial(datos["usd"])
        bloque("Instrumentros a licitar en dólares",datos["usd"],es_usd=True)
        if any(x.get("tipo")=="BONAR" for x in datos["usd"]):
            r=R(2)
            ws.cell(row=r,column=CL,
                value="(*) En segunda vuelta se emitirá un monto tal que en conjunto con la primera vuelta no supere el VNO USD de 250 Millones"
            ).font=Font(size=SZ,name="Calibri")

    # ── Bloque canje ──────────────────────────────────────────────────────────
    if datos.get("canje"):
        R(2)
        titulo_canje = f"Licitación para la Conversión del {datos.get('titulo_elegible','')}:"
        liq_str = datos.get("liq_canje_str","")
        r=R(); ws.row_dimensions[r].height=21
        ct=ws.cell(row=r,column=CL,value=titulo_canje)
        ct.font=Font(italic=True,bold=True,size=SZ,name="Calibri"); ct.alignment=al(h="left",wrap=False)
        if liq_str:
            r=R(); ws.row_dimensions[r].height=18
            c=ws.cell(row=r,column=CL,value=f"Liquidación: {liq_str} (T+3)")
            c.font=Font(size=SZ,name="Calibri"); c.alignment=al(h="left",wrap=False)

        canje_insts = datos["canje"]
        n=len(canje_insts); c1=CD; c2=CD+n-1
        R()  # vacía

        # Fila especial: Nombre del Título Elegible (combinada)
        r=R(); ws.row_dimensions[r].height=42
        cc=ws.cell(row=r,column=CL); cc.fill=F(NARANJA); cc.border=B()
        for i,inst in enumerate(canje_insts):
            c=ws.cell(row=r,column=c1+i,value=inst["label"])
            c.fill=F(NARANJA); c.font=fn(color=BLANCO)
            c.alignment=al(); c.border=B()

        # Fila Nombre del Título Elegible
        r=R(); ws.row_dimensions[r].height=21
        aplicar(ws.cell(row=r,column=CL),GRIS_OSC,valor="Nombre del Título Elegible",h="left")
        merge_escribir(ws,r,c1,c2,GRIS_OSC,datos.get("titulo_elegible",""))

        # Filas estándar
        def fi(label,getter,gris,h_row=21):
            r=R(); ws.row_dimensions[r].height=h_row
            aplicar(ws.cell(row=r,column=CL),gris,valor=label,h="left")
            for i,inst in enumerate(canje_insts):
                aplicar(ws.cell(row=r,column=c1+i),gris,valor=getter(inst) or "")

        def fm(label,valor,gris,h_row=21):
            r=R(); ws.row_dimensions[r].height=h_row
            aplicar(ws.cell(row=r,column=CL),gris,valor=label,h="left")
            merge_escribir(ws,r,c1,c2,gris,valor)

        fi("Vencimiento",lambda x:x["vencimiento"].strftime("%d/%m/%Y") if x["vencimiento"] else "",GRIS_CLA)
        fi("Tasa de interés",lambda x:x.get("tasa",""),GRIS_OSC)
        fi("Precio",lambda x:x.get("precio",""),GRIS_CLA,h_row=42)
        fi("Ajuste de capital",lambda x:x.get("ajuste",""),GRIS_OSC)
        fi("Párametro a licitar",lambda x:x.get("parametro",""),GRIS_CLA)
        fm("Amortización","Íntegra al vencimiento",GRIS_OSC)
        fm("Monto Máximo a Licitar","Hasta el monto máximo autorizado por la normativa vigente",GRIS_CLA)
        fm("Ley aplicable","Ley de la REPÚBLICA ARGENTINA",GRIS_OSC)

    out=io.BytesIO(); wb.save(out); out.seek(0)
    return out

# ════════════════════════════════════════════════════════════════════════════
# MÓDULO RESULTADOS
# ════════════════════════════════════════════════════════════════════════════

NUM = r"[\d\.]+(?:,\d+)?"  # número: dígitos con punto miles y coma decimal

def _label_resultado(ticker, tipo_ticker, tipo_bloque, nombre_raw):
    """Genera label corto para la columna de resultado."""
    es_nuevo = "NUEVO" in tipo_ticker.upper()
    if tipo_bloque == "pesos_fija":
        if "BONO" in nombre_raw and "CAPITALIZABLE" in nombre_raw:
            return f"BONCAP ({ticker} - reapertura)"
        return f"LECAP ({ticker} - reapertura)"
    elif tipo_bloque == "cer":
        return f"BONCER ({ticker} – reapertura)"
    elif tipo_bloque == "tamar":
        if "LETRA" in nombre_raw:
            return f"LETAM ({ticker} - reapertura)"
        return f"BOTAM ({ticker} - nuevo)" if es_nuevo else f"BOTAM ({ticker} - reapertura)"
    elif tipo_bloque == "usd":
        return f"LELINK ({ticker} - reapertura)"
    return ticker

def _extraer_pesos(seccion, tipo_bloque):
    """Extrae instrumentos de secciones pesos (tasa fija y CER)."""
    pat = (
        r"\$\s*(" + NUM + r")\s+"
        r"\$\s*(" + NUM + r")\s+"
        r"\$\s*(" + NUM + r")\s+"
        r"\$\s*(" + NUM + r")\s+"
        r"(" + NUM + r")%\s+"
        r"\$\s*(" + NUM + r")\s+"
        r"(?:[\w\s]+?\s+)?"
        r"\((\w+)\s*[-–]\s*(REAPERTURA|NUEVO)\)"
    )
    res = []
    for m in re.finditer(pat, seccion):
        pos = m.start()
        antes = seccion[:pos].strip()
        partes = re.split(r"\)\s*(?:\(\d+\))?\s*", antes)
        nombre = partes[-1].strip()
        nombre = re.sub(r"\s*\(\d+\)\s*$", "", nombre).strip()
        ticker = m.group(7); tipo_t = m.group(8)
        res.append({
            "label": _label_resultado(ticker, tipo_t, tipo_bloque, nombre),
            "vno_ofertado":  f"$ {m.group(1)}",
            "vno_adjudicado": f"$ {m.group(2)}",
            "ve_adjudicado":  f"$ {m.group(3)}",
            "precio_corte":   f"$ {m.group(4)}",
            "tirea":          f"{m.group(5)}%",
            "vno_circulacion": f"$ {m.group(6)}",
        })
    return res

def _extraer_tamar(seccion):
    """Extrae instrumentos TAMAR (LETAM con precio $, BOTAM nuevo con margen %)."""
    res = []
    # LETAM reapertura: precio $
    pat_letam = (
        r"\$\s*(" + NUM + r")\s+"
        r"\$\s*(" + NUM + r")\s+"
        r"\$\s*(" + NUM + r")\s+"
        r"\$\s*(" + NUM + r")\s+"
        r"(" + NUM + r")%\s+"
        r"\$\s*(" + NUM + r")\s+"
        r"(?:[\w\s]+?\s+)?"
        r"\((\w+)\s*-\s*(REAPERTURA)\)"
    )
    for m in re.finditer(pat_letam, seccion):
        pos = m.start()
        nombre = seccion[:pos].strip().split(")")[-1].strip()
        ticker = m.group(7)
        res.append({
            "label": _label_resultado(ticker, "REAPERTURA", "tamar", nombre),
            "vno_ofertado":   f"$ {m.group(1)}",
            "vno_adjudicado": f"$ {m.group(2)}",
            "ve_adjudicado":  f"$ {m.group(3)}",
            "precio_corte":   f"$ {m.group(4)}",
            "tirea":          f"{m.group(5)}%",
            "vno_circulacion": f"$ {m.group(6)}",
        })
    # BOTAM nuevo: margen % en lugar de precio $
    pat_botam = (
        r"\$\s*(" + NUM + r")\s+"
        r"\$\s*(" + NUM + r")\s+"
        r"\$\s*(" + NUM + r")\s+"
        r"(" + NUM + r")%\s+"   # margen de corte
        r"(" + NUM + r")%\s+"   # tirea
        r"\$\s*(" + NUM + r")\s+"
        r"(?:[\w\s]+?\s+)?"
        r"\((\w+)\s*[-–]?\s*(NUEVO)\)"
    )
    for m in re.finditer(pat_botam, seccion):
        pos = m.start()
        nombre = seccion[:pos].strip().split(")")[-1].strip()
        ticker = m.group(7)
        res.append({
            "label": _label_resultado(ticker, "NUEVO", "tamar", nombre),
            "vno_ofertado":   f"$ {m.group(1)}",
            "vno_adjudicado": f"$ {m.group(2)}",
            "ve_adjudicado":  f"$ {m.group(3)}",
            "precio_corte":   f"{m.group(4)}%",  # margen
            "tirea":          f"{m.group(5)}%",
            "vno_circulacion": f"$ {m.group(6)}",
        })
    # Ordenar por posición en el texto
    res.sort(key=lambda x: seccion.find(x["vno_ofertado"].replace("$ ","")))
    return res

def _extraer_usd(seccion):
    """Extrae instrumentos USD: VNO en USD, VE en $."""
    pat = (
        r"USD\s*(" + NUM + r")\s+"
        r"USD\s*(" + NUM + r")\s+"
        r"\$\s*(" + NUM + r")\s+"
        r"USD\s*(" + NUM + r")\s+"
        r"(" + NUM + r")%\s+"
        r"USD\s*(" + NUM + r")\s+"
        r"(?:[\w\s]+?\s+)?"
        r"\((\w+)\s*-\s*(REAPERTURA)\)"
    )
    res = []
    for m in re.finditer(pat, seccion):
        pos = m.start()
        nombre = seccion[:pos].strip().split(")")[-1].strip()
        ticker = m.group(7)
        res.append({
            "label": _label_resultado(ticker, "REAPERTURA", "usd", nombre),
            "vno_ofertado":   f"USD {m.group(1)}",
            "vno_adjudicado": f"USD {m.group(2)}",
            "ve_adjudicado":  f"$ {m.group(3)}",
            "precio_corte":   f"USD {m.group(4)}",
            "tirea":          f"{m.group(5)}%",
            "vno_circulacion": f"USD {m.group(6)}",
        })
    return res

def extraer_resultados_pdf(pdf_bytes):
    """Extrae todos los datos del PDF de resultados."""
    txt = ""
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for p in pdf.pages: txt += (p.extract_text() or "") + "\n"
    T = re.sub(r"\s+", " ", txt.upper()).strip()

    # Fecha de la licitación (del pie del documento)
    mf = re.search(r"BUENOS AIRES.*?(\d{1,2}\s+DE\s+\w+\s+DE\s+\d{4})", T)
    fecha_lic = pf(mf.group(1)) if mf else None

    # Fecha liquidación: buscar explícita, si no calcular T+2 hábiles desde fecha_lic
    m_liq = re.search(r"LIQUIDACI[ÓO]N.*?(\d{1,2}\s+DE\s+\w+\s+DE\s+\d{4})", T)
    if m_liq:
        fecha_liq = pf(m_liq.group(1))
    elif fecha_lic:
        # Calcular T+2 hábiles (saltar fines de semana)
        from datetime import timedelta
        f = fecha_lic
        dias_habiles = 0
        while dias_habiles < 2:
            f = f + timedelta(days=1)
            if f.weekday() < 5:  # lunes a viernes
                dias_habiles += 1
        fecha_liq = f
    else:
        fecha_liq = None

    # Tipo de cambio
    m_tc = re.search(r"PESOS\s*/\s*USD\s*([\d\.\,]+)", T)
    tc_str = m_tc.group(1) if m_tc else ""
    m_tc_f = re.search(r"D[ÍI]A\s+(\d{1,2}\s+DE\s+\w+\s+DE\s+\d{4})", T)
    tc_fecha = m_tc_f.group(1).strip().title() if m_tc_f else ""

    # Totales
    totales = {}
    for campo, pat in [
        ("ofertas",    r"CANTIDAD DE OFERTAS RECIBIDAS\s+([\d\.]+)\s+([\d\.]+)\s+([\d\.]+)"),
        ("vno_of",     r"TOTAL VNO OFERTADO[^$]*\$\s*([\d\.]+)\s+USD\s*([\d]+)\s*\$\s*([\d\.]+)"),
        ("vno_adj",    r"TOTAL VNO ADJUDICADO[^$]*\$\s*([\d\.]+)\s+USD\s*([\d]+)\s*\$\s*([\d\.]+)"),
        ("ve_adj",     r"TOTAL VALOR EFECTIVO ADJUDICADO[^$]*\$\s*([\d\.]+)\s*\$\s*([\d\.]+)\s*\$\s*([\d\.]+)"),
    ]:
        m = re.search(pat, T)
        if m: totales[campo] = (m.group(1), m.group(2), m.group(3))

    # Extraer bloques de instrumentos
    SECCIONES = [
        ("pesos_fija",
         "Instrumentos licitados denominados en pesos a tasa fija:",
         "INSTRUMENTOS LICITADOS DENOMINADOS EN PESOS A TASA FIJA",
         "INSTRUMENTOS LICITADOS DENOMINADOS EN PESOS CON AJUSTE",
         lambda s: _extraer_pesos(s, "pesos_fija")),
        ("cer",
         "Instrumentos licitados denominados en pesos ajustados por CER:",
         "INSTRUMENTOS LICITADOS DENOMINADOS EN PESOS CON AJUSTE POR CER",
         "INSTRUMENTOS LICITADOS DENOMINADOS EN PESOS A TASA TAMAR",
         lambda s: _extraer_pesos(s, "cer")),
        ("tamar",
         "Instrumentos licitados denominados en pesos a tasa TAMAR:",
         "INSTRUMENTOS LICITADOS DENOMINADOS EN PESOS A TASA TAMAR",
         "INSTRUMENTO LICITADO DENOMINADO EN D",
         _extraer_tamar),
        ("usd",
         "Instrumento licitado denominado en dólares estadounidenses:",
         "INSTRUMENTO LICITADO DENOMINADO EN D",
         "BUENOS AIRES",
         _extraer_usd),
    ]

    bloques = []
    for tipo, titulo, pat_ini, pat_fin, fn in SECCIONES:
        m_ini = re.search(pat_ini, T)
        m_fin = re.search(pat_fin, T)
        if not m_ini: continue
        start = m_ini.end()
        end = m_fin.start() if m_fin and m_fin.start() > start else len(T)
        insts = fn(T[start:end])
        if insts:
            bloques.append({"tipo": tipo, "titulo": titulo, "instrumentos": insts})

    return {
        "bloques": bloques, "totales": totales,
        "fecha_liq": fecha_liq, "fecha_liq_str": fstr(fecha_liq),
        "fecha_lic": fecha_lic, "tc_str": tc_str, "tc_fecha": tc_fecha,
    }


def generar_excel_resultados(datos):
    """Genera el Excel de resultados con el mismo formato visual que el de llamado."""
    wb = Workbook(); ws = wb.active
    ws.sheet_view.showGridLines = False
    fecha_lic = datos["fecha_lic"]
    if fecha_lic: ws.title = f"Rdo {fecha_lic.strftime('%d.%m.%y')}"

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 4
    ws.column_dimensions["C"].width = 4
    ws.column_dimensions["D"].width = 4
    ws.column_dimensions["E"].width = 32
    for i in range(6, 20): ws.column_dimensions[get_column_letter(i)].width = 22

    # Logos
    ws.row_dimensions[2].height = 50
    try:
        buf, _, _ = redimensionar_logo(LOGO_HIP_PATH, 50)
        img = XLImage(buf); img.anchor = "C2"; ws.add_image(img)
    except: pass
    try:
        buf, _, _ = redimensionar_logo(LOGO_MIN_PATH, 50)
        img = XLImage(buf); img.anchor = "H2"; ws.add_image(img)
    except: pass

    # Header
    ws.row_dimensions[9].height = 18
    c = ws.cell(row=9, column=7, value="LICITACIÓN DEL TESORO")
    c.font = Font(bold=True, size=SZ, name="Calibri"); c.alignment = al(h="center", wrap=False)

    ws.row_dimensions[11].height = 18
    c = ws.cell(row=11, column=5, value="Resultados de la licitación por Efectivo:")
    c.font = Font(bold=True, size=SZ, name="Calibri"); c.alignment = al(h="left", wrap=False)
    ws.merge_cells(start_row=11, start_column=5, end_row=11, end_column=14)

    ws.row_dimensions[13].height = 18
    c = ws.cell(row=13, column=5, value=f"Fecha de Liquidación: {datos['fecha_liq_str']} (T+2)")
    c.font = Font(size=SZ, name="Calibri"); c.alignment = al(h="left", wrap=False)
    ws.merge_cells(start_row=13, start_column=5, end_row=13, end_column=14)

    fila = [15]
    def R(n=1): r = fila[0]; fila[0] += n; return r
    CL = 5; CD = 6  # col labels = E, datos desde F

    FILAS_DATOS = [
        ("VN Ofertado (*)",             "vno_ofertado",   GRIS_OSC),
        ("VN Adjudicado (*)",           "vno_adjudicado", GRIS_CLA),
        ("Valor Efectivo Adjudicado (*)", "ve_adjudicado",  GRIS_OSC),
        ("Precio/Tasa de Corte ",       "precio_corte",   GRIS_CLA),
        ("TIREA",                       "tirea",          GRIS_OSC),
        ("VNO total circulación (*)",   "vno_circulacion", GRIS_CLA),
    ]

    for bloque in datos["bloques"]:
        insts = bloque["instrumentos"]
        if not insts: continue
        n = len(insts); c1 = CD; c2 = CD + n - 1

        # Título sección
        r = R(); ws.row_dimensions[r].height = 21
        ct = ws.cell(row=r, column=CL, value=bloque["titulo"])
        ct.font = Font(italic=True, bold=True, size=SZ, name="Calibri")
        ct.alignment = al(h="left", wrap=False)
        ws.merge_cells(start_row=r, start_column=CL, end_row=r, end_column=c2)
        R()  # vacía

        # Header naranja
        r = R(); ws.row_dimensions[r].height = 42
        ws.cell(row=r, column=CL).fill = F(NARANJA); ws.cell(row=r, column=CL).border = B()
        for i, inst in enumerate(insts):
            c = ws.cell(row=r, column=c1+i, value=inst["label"])
            c.fill = F(NARANJA); c.font = fn(color=BLANCO)
            c.alignment = al(); c.border = B()

        # Filas de datos
        for label, campo, gris in FILAS_DATOS:
            r = R(); ws.row_dimensions[r].height = 21
            aplicar(ws.cell(row=r, column=CL), gris, valor=label, h="left")
            for i, inst in enumerate(insts):
                val = inst.get(campo, "") if campo else ""
                aplicar(ws.cell(row=r, column=c1+i), gris, valor=val)

        R()  # vacía

    # Tabla resumen totales
    tot = datos["totales"]
    if tot:
        r = R(); ws.row_dimensions[r].height = 21
        ct = ws.cell(row=r, column=CL, value="Resumen de la Licitación")
        ct.font = Font(italic=True, bold=True, size=SZ, name="Calibri")
        ct.alignment = al(h="left", wrap=False)
        R()

        # Headers resumen
        r = R(); ws.row_dimensions[r].height = 42
        hdrs = ["", "Resultados en pesos", "Resultado en dólares", "TOTAL"]
        for i, txt_h in enumerate(hdrs):
            c = ws.cell(row=r, column=CL+i, value=txt_h)
            c.fill = F(NARANJA); c.font = fn(color=BLANCO)
            c.alignment = al(); c.border = B()

        FILAS_TOT = [
            ("Cantidad de Ofertas Recibidas", "ofertas", GRIS_OSC),
            ("Total VNO Ofertado (*)",        "vno_of",  GRIS_CLA),
            ("Total VNO Adjudicado (*)",      "vno_adj", GRIS_OSC),
            ("Total Valor Efectivo Adjudicado (*)", "ve_adj", GRIS_CLA),
        ]
        for label, campo, gris in FILAS_TOT:
            r = R(); ws.row_dimensions[r].height = 21
            aplicar(ws.cell(row=r, column=CL), gris, valor=label, h="left")
            d = tot.get(campo, ("","",""))
            aplicar(ws.cell(row=r, column=CL+1), gris, valor=f"$ {d[0]}" if d[0] else "")
            aplicar(ws.cell(row=r, column=CL+2), gris, valor=f"USD {d[1]}" if d[1] else "")
            aplicar(ws.cell(row=r, column=CL+3), gris, valor=f"$ {d[2]}" if d[2] else "")

        R()
        r = R(); ws.row_dimensions[r].height = 21
        nota = "(*) Montos expresados en millones."
        if datos["tc_str"]:
            nota += f"  (**) Tipo de Cambio de Referencia {datos['tc_fecha']}: Pesos / USD {datos['tc_str']}."
        c = ws.cell(row=r, column=CL, value=nota)
        c.font = Font(size=9, italic=True, name="Calibri")
        c.alignment = al(h="left", wrap=True)
        ws.merge_cells(start_row=r, start_column=CL, end_row=r, end_column=CL+5)

    out = io.BytesIO(); wb.save(out); out.seek(0)
    return out

# ── UI ────────────────────────────────────────────────────────────────────────
st.markdown("## 🏦 Licitación del Tesoro")
st.markdown("**Banco Hipotecario · Mercado de Capitales**")
st.markdown("---")

tab1, tab2 = st.tabs(["📋 Llamado", "📊 Resultados"])

with tab1:
    st.markdown("### Subí el PDF del nuevo llamado")
    pdf_file = st.file_uploader("📄 PDF del llamado", type=["pdf"], key="llamado")
    st.markdown("---")
    if pdf_file:
        if st.button("⚙️ Generar Excel del Llamado"):
            with st.spinner("Procesando..."):
                try:
                    datos = extraer_pdf(pdf_file.read())
                    excel_out = generar_excel(datos)
                    total = sum(len(datos[k]) for k in ["pf","pv","cer","usd"])
                    st.success(f"✅ {total} instrumentos detectados")
                    for bq, nm in [("pf","💵 Tasa fija"),("pv","📊 Tasa variable"),("cer","📈 CER"),("usd","💲 Dólares")]:
                        if datos[bq]:
                            st.markdown(f"**{nm}**")
                            for inst in datos[bq]:
                                v = inst["vencimiento"].strftime("%d/%m/%Y") if inst["vencimiento"] else "N/A"
                                st.markdown(f"&nbsp;&nbsp;&nbsp;• {inst.get("label", "")} — vto. {v}")
                    if datos["h"]["liq_str"]:
                        st.markdown(f"📅 **Liquidación:** {datos['h']['liq_str']}")
                    fl = datos["h"]["liq"]
                    nombre = f"Licitacion_Tesoro_{fl.strftime('%d_%m_%Y')}.xlsx" if fl else "Licitacion_nueva.xlsx"
                    st.markdown("---")
                    st.download_button("⬇️ Descargar Excel", data=excel_out, file_name=nombre,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                except Exception as e:
                    st.error(f"❌ Error: {str(e)}"); st.exception(e)
    else:
        st.info("⬆️ Subí el PDF para habilitar la generación.")

with tab2:
    st.markdown("### Subí el PDF de resultados")
    pdf_res = st.file_uploader("📄 PDF de resultados", type=["pdf"], key="resultados")
    st.markdown("---")
    if pdf_res:
        if st.button("⚙️ Generar Excel de Resultados"):
            with st.spinner("Procesando resultados..."):
                try:
                    datos_res = extraer_resultados_pdf(pdf_res.read())
                    excel_res = generar_excel_resultados(datos_res)
                    total_insts = sum(len(b["instrumentos"]) for b in datos_res["bloques"])
                    st.success(f"✅ {total_insts} instrumentos detectados")
                    for b in datos_res["bloques"]:
                        st.markdown(f"**{b['titulo']}**")
                        for inst in b["instrumentos"]:
                            st.markdown(f"&nbsp;&nbsp;&nbsp;• {inst['label']}")
                    if datos_res["fecha_liq_str"]:
                        st.markdown(f"📅 **Liquidación:** {datos_res['fecha_liq_str']}")
                    fl = datos_res["fecha_lic"]
                    nombre_res = f"Resultado_Tesoro_{fl.strftime('%d_%m_%Y')}.xlsx" if fl else "Resultado_nueva.xlsx"
                    st.markdown("---")
                    st.download_button("⬇️ Descargar Excel de Resultados", data=excel_res, file_name=nombre_res,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                except Exception as e:
                    st.error(f"❌ Error: {str(e)}"); st.exception(e)
    else:
        st.info("⬆️ Subí el PDF de resultados para habilitar la generación.")

st.markdown("---")
st.caption("Banco Hipotecario · Mercado de Capitales · Emisiones Primarias")




